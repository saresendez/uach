import os
import re
import io
import sys
import json
import time
import base64
import requests
import datetime
import email.utils
import pandas as pd

# Librerías oficiales de Google para OAuth2, Gmail y Drive APIs
from google.auth.transport.requests import Request
from google.oauth2.credentials import Credentials
from google_auth_oauthlib.flow import Flow
from googleapiclient.discovery import build
from googleapiclient.http import MediaIoBaseDownload, MediaFileUpload

# Librerías para dar formato avanzado a celdas de Excel (Color, Negrita, Cursiva)
import openpyxl
from openpyxl.styles import Font

# ==========================================
# CONFIGURACIÓN Y CONSTANTES
# ==========================================
MODELO_IA = "gemini-2.5-flash"

# Nombres de archivos de forma local en tu VPS
ARCHIVO_EXCEL_LOCAL = "reg_anon.xlsx"
ARCHIVO_MAPEO_LOCAL = "mapeo_identidad.xlsx"
ARCHIVO_REGISTROS_LOCAL = "registros.xlsx"

# Nombres exactos de tus archivos en Google Drive
NOMBRE_EXCEL_DRIVE = "reg_anon.xlsx"
NOMBRE_MAPEO_DRIVE = "mapeo_identidad.xlsx"
NOMBRE_REGISTROS_DRIVE = "registros.xlsx"

# Rutas locales para credenciales oficiales en tu VPS o entorno local
RUTA_CREDENTIALS = "credentials.json"
RUTA_TOKEN = "token.json"

# Nombre de la etiqueta de control para evitar duplicados en Gmail
ETIQUETA_CONTROL = "CONCILIADO"

# Permisos requeridos para leer/modificar correos de Gmail y escribir en Drive
SCOPES = [
    'https://www.googleapis.com/auth/gmail.modify',
    'https://www.googleapis.com/auth/drive'
]

# Permitir que oauthlib acepte URLs HTTP locales de redirección
os.environ['OAUTHLIB_INSECURE_TRANSPORT'] = '1'
os.environ['OAUTHLIB_RELAX_TOKEN_SCOPE'] = '1'

# Meses del año en español para cascada
MESES = ["ENERO", "FEBRERO", "MARZO", "ABRIL", "MAYO", "JUNIO", "JULIO", "AGOSTO", "SEPTIEMBRE", "OCTUBRE", "NOVIEMBRE", "DICIEMBRE"]
PALABRAS_CLAVE_PAGO = ["PAGO", "COMPROBANTE", "TRANSFERENCIA", "RECIBO", "DEPÓSITO", "DEPOSITO"]

# Recuperación segura de API Key
API_KEY = os.environ.get('GEMINI_API_KEY')

# ==========================================
# AUXILIARES DE ENTORNO Y FORMATEO
# ==========================================
def normalizar_id_cliente(valor):
    """Normaliza IDs de cliente para evitar discrepancias de tipo de datos (ej. 102 vs 102.0)."""
    if pd.isna(valor):
        return ""
    val_str = str(valor).strip().upper()
    if val_str.endswith('.0'):
        val_str = val_str[:-2]
    return val_str

# ==========================================
# AUTENTICACIÓN HÍBRIDA (OAUTH 2.0)
# ==========================================
def obtener_credenciales():
    """Recupera o solicita credenciales OAuth2 de forma híbrida."""
    creds = None
    if os.path.exists(RUTA_TOKEN):
        creds = Credentials.from_authorized_user_file(RUTA_TOKEN, SCOPES)

    if not creds or not creds.valid:
        if creds and creds.expired and creds.refresh_token:
            try:
                creds.refresh(Request())
            except Exception as e:
                print(f"⚠️ El token expiró y no pudo renovarse de forma automática: {e}. Requiriendo reautorización.")
                creds = None
        
        if not creds:
            if not os.path.exists(RUTA_CREDENTIALS):
                raise FileNotFoundError(f"❌ Error crítico: Falta el archivo '{RUTA_CREDENTIALS}' en tu espacio de trabajo.")

            flow = Flow.from_client_secrets_file(
                RUTA_CREDENTIALS,
                scopes=SCOPES,
                redirect_uri='http://localhost'
            )

            authorization_url, state = flow.authorization_url(
                access_type='offline',
                include_granted_scopes='true'
            )

            print("\n================ AUTORIZACIÓN REQUERIDA ================")
            print("Abre este enlace en tu navegador para autorizar la aplicación:")
            print(f"\n👉 {authorization_url}\n")
            print("Copia la URL completa de localhost (ej. http://localhost/?code=...) al finalizar:")
            print("========================================================\n")

            url_redirigida = input("Pega la URL de localhost aquí: ").strip()
            flow.fetch_token(authorization_response=url_redirigida)
            creds = flow.credentials

        # Guardar credenciales para futuras ejecuciones automáticas (desatendidas)
        with open(RUTA_TOKEN, 'w') as token_file:
            token_file.write(creds.to_json())

    return creds

# ==========================================
# INTEGRACIÓN EXCLUSIVA CON GOOGLE DRIVE API
# ==========================================
def descargar_archivo_de_drive(servicio_drive, nombre_archivo, ruta_destino):
    """Busca un archivo por nombre en tu Google Drive y lo descarga localmente."""
    try:
        query = f"name = '{nombre_archivo}' and trashed = false"
        resultados = servicio_drive.files().list(q=query, fields="files(id, name)").execute()
        archivos = resultados.get('files', [])

        if not archivos:
            print(f"❌ No se encontró el archivo '{nombre_archivo}' en Google Drive.")
            return None

        file_id = archivos[0]['id']
        request = servicio_drive.files().get_media(fileId=file_id)
        
        fh = io.BytesIO()
        downloader = MediaIoBaseDownload(fh, request)
        done = False
        while not done:
            status, done = downloader.next_chunk()

        fh.seek(0)
        with open(ruta_destino, 'wb') as f:
            f.write(fh.read())
        print(f"📥 Archivo '{nombre_archivo}' descargado de Google Drive con éxito.")
        return file_id
    except Exception as e:
        print(f"❌ Error al descargar '{nombre_archivo}' de Drive: {e}")
        return None

def actualizar_archivo_en_drive(servicio_drive, file_id, ruta_local):
    """Actualiza los datos del archivo en Google Drive de forma remota sin crear duplicados."""
    try:
        media = MediaFileUpload(
            ruta_local, 
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', 
            resumable=True
        )
        servicio_drive.files().update(fileId=file_id, media_body=media).execute()
        print(f"📤 ¡Cambios sincronizados exitosamente con Google Drive! (ID: {file_id})")
        return True
    except Exception as e:
        print(f"❌ Error al sincronizar cambios con Google Drive: {e}")
        return False

# ==========================================
# GESTIÓN DE ETIQUETAS EN GMAIL
# ==========================================
def obtener_o_crear_etiqueta(servicio_gmail, nombre_etiqueta):
    try:
        resultados = servicio_gmail.users().labels().list(userId='me').execute()
        etiquetas = resultados.get('labels', [])

        for etiqueta in etiquetas:
            if etiqueta['name'].upper() == nombre_etiqueta.upper():
                return etiqueta['id']

        print(f"🏷️ Creando etiqueta '{nombre_etiqueta}' en Gmail...")
        nueva_etiqueta_body = {
            "name": nombre_etiqueta,
            "labelListVisibility": "labelShow",
            "messageListVisibility": "show"
        }
        creada = servicio_gmail.users().labels().create(userId='me', body=nueva_etiqueta_body).execute()
        return creada['id']
    except Exception as e:
        print(f"⚠️ Error al gestionar etiquetas de Gmail: {e}")
        return None

def aplicar_etiqueta_procesado(servicio_gmail, message_id, label_id):
    try:
        servicio_gmail.users().messages().batchModify(
            userId='me',
            body={
                'ids': [message_id],
                'addLabelIds': [label_id],
                'removeLabelIds': ['UNREAD']
            }
        ).execute()
        print(f"    🏷️ Correo {message_id} marcado como '{ETIQUETA_CONTROL}' y leído.")
    except Exception as e:
        print(f"    ❌ No se pudo aplicar etiqueta al correo {message_id}: {e}")

# ==========================================
# PROCESAMIENTO DE TEXTO DE CORREOS
# ==========================================
def extraer_cuerpo_texto(payload):
    """Extrae recursivamente el contenido de texto (plano o HTML) del correo."""
    def _extraer(parte):
        mime_type = parte.get('mimeType', '')
        body_data = parte.get('body', {}).get('data', '')
        texto = ""

        if mime_type == 'text/plain' and body_data:
            try:
                texto = base64.urlsafe_b64decode(body_data).decode('utf-8', errors='ignore')
            except Exception:
                pass
        elif mime_type == 'text/html' and body_data:
            try:
                html_raw = base64.urlsafe_b64decode(body_data).decode('utf-8', errors='ignore')
                texto = re.sub(r'<[^>]+>', ' ', html_raw)
                texto = texto.replace('&nbsp;', ' ').replace('&amp;', '&').replace('&gt;', '>')
            except Exception:
                pass

        partes = parte.get('parts', [])
        texto_acumulado = texto
        for p in partes:
            texto_acumulado += " " + _extraer(p)
        return texto_acumulado

    cuerpo_sucio = _extraer(payload)
    return re.sub(r'\s+', ' ', cuerpo_sucio).strip()

def buscar_palabras_clave(asunto, cuerpo):
    """Verifica si en el asunto o cuerpo aparece alguna palabra clave de pago."""
    texto_completo = f"{asunto} {cuerpo}".upper()
    return any(palabra in texto_completo for palabra in PALABRAS_CLAVE_PAGO)

def detectar_mes_en_texto(asunto, cuerpo):
    """Escanea el asunto y el cuerpo del correo buscando menciones de algún mes."""
    texto_completo = f" {asunto} {cuerpo} ".upper()
    texto_limpio = re.sub(r'[.,;:()\-_\t/]', ' ', texto_completo)
    
    for mes in MESES:
        patron = r'\b' + mes + r'\b'
        if re.search(patron, texto_limpio):
            return mes
    return None

def detectar_mes_en_hilo(servicio_gmail, thread_id):
    """Obtiene todos los mensajes del hilo de conversación para escanear el historial completo."""
    try:
        thread = servicio_gmail.users().threads().get(userId='me', id=thread_id).execute()
        mensajes_hilo = thread.get('messages', [])
        
        texto_acumulado_hilo = ""
        for m in mensajes_hilo:
            p = m.get('payload', {})
            headers = p.get('headers', [])
            asunto = next((h.get('value', '') for h in headers if h.get('name', '').lower() == 'subject'), "")
            cuerpo = extraer_cuerpo_texto(p)
            texto_acumulado_hilo += f" {asunto} {cuerpo}"
            
        return detectar_mes_en_texto("", texto_acumulado_hilo)
    except Exception as e:
        print(f"   ⚠️ No se pudo escanear el historial del hilo de conversación: {e}")
    return None

def obtener_mes_desde_fecha(fecha_str):
    """Extrae el mes en español a partir de una fecha con formato DD/MM/YYYY."""
    try:
        partes = fecha_str.split('/')
        if len(partes) == 3:
            mes_idx = int(partes[1])
            if 1 <= mes_idx <= 12:
                return MESES[mes_idx - 1]
    except Exception:
        pass
    return None

def obtener_adjuntos_recursivo(parte):
    adjuntos = []
    filename = parte.get('filename')
    body = parte.get('body', {})
    attachment_id = body.get('attachmentId')

    if filename and attachment_id:
        adjuntos.append(parte)

    partes_hijas = parte.get('parts', [])
    for p in partes_hijas:
        adjuntos.extend(obtener_adjuntos_recursivo(p))
    return adjuntos

# ==========================================
# PROCESAMIENTO MULTIMODAL CON GEMINI
# ==========================================
def limpiar_y_cargar_json(texto_respuesta):
    try:
        texto_limpio = re.sub(r'^\s*```(?:json)?\s*', '', texto_respuesta, flags=re.MULTILINE | re.IGNORECASE)
        texto_limpio = re.sub(r'\s*```$', '', texto_limpio, flags=re.MULTILINE)
        return json.loads(texto_limpio.strip())
    except Exception as e:
        print(f"    ❌ Error al decodificar JSON de la IA: {e}")
        return None

def extraer_metadatos_con_gemini(contenido_binario, mime_type):
    if not API_KEY:
        print("⚠️ Advertencia: API_KEY no configurada. Saltando análisis con IA.")
        return None

    url = f"https://generativelanguage.googleapis.com/v1beta/models/{MODELO_IA}:generateContent?key={API_KEY}"
    try:
        archivo_base64 = base64.b64encode(contenido_binario).decode('utf-8')
        system_prompt = (
            "Actúa como un extractor de metadatos estructurados especializado en recibos de pago. "
            "Analiza el archivo adjunto y realiza lo siguiente:\n"
            "1. Si el archivo adjunto es un comprobante de pago, transferencia bancaria o recibo de depósito válido, "
            "extrae la fecha del pago (en formato string 'DD/MM/YYYY') y el monto total pagado como número flotante sin símbolos de moneda. "
            "2. Si el archivo adjunto NO es un comprobante de pago o transferencia bancaria válido, responde estrictamente con valores nulos.\n\n"
            "Responde ÚNICAMENTE con este JSON exacto:\n"
            '{"fecha": "DD/MM/YYYY", "monto": 0.0}'
        )
        payload = {
            "contents": [{"parts": [
                {"text": system_prompt},
                {"inlineData": {"mimeType": mime_type, "data": archivo_base64}}
            ]}],
            "generationConfig": {"responseMimeType": "application/json"}
        }
        
        # Exponential Backoff robusto (5 intentos)
        delays = [2, 4, 8, 16, 32]
        for intento, delay in enumerate(delays):
            response = requests.post(url, headers={'Content-Type': 'application/json'}, json=payload)
            if response.status_code == 200:
                texto_respuesta = response.json()['candidates'][0]['content']['parts'][0]['text'].strip()
                # Pausa preventiva de 4 segundos para respetar el límite de 15 RPM del plan gratuito
                time.sleep(4)
                return limpiar_y_cargar_json(texto_respuesta)
            elif response.status_code in [429, 503]:
                print(f"   ⏳ [Intento {intento+1}/{len(delays)}] Cuota excedida o servidor saturado ({response.status_code}). Reintentando en {delay}s...")
                time.sleep(delay)
            else:
                print(f"   ⚠️ Error de API inesperado (Código {response.status_code}). Deteniendo reintentos.")
                break
        return None
    except Exception as e:
        print(f"    ❌ Error en Gemini: {e}")
        return None

# ==========================================
# ACTUALIZACIÓN EN EXCEL CON ESTILOS AVANZADOS
# ==========================================
def actualizar_excel(id_cliente, mes_pago, monto_pago):
    """
    Registra el pago en el Excel. Si el mes seleccionado ya está liquidado (tiene datos),
    aplica el monto al siguiente mes disponible en orden cronológico, pintando la celda
    con texto en VERDE, NEGRITA y CURSIVA usando openpyxl.
    """
    try:
        # 1. ACTUALIZACIÓN EN reg_anon.xlsx
        wb = openpyxl.load_workbook(ARCHIVO_EXCEL_LOCAL)
        ws = wb.active

        # Mapear los encabezados (Columnas)
        encabezados = [str(ws.cell(row=1, column=col).value).strip().upper() for col in range(1, ws.max_column + 1)]
        
        id_cliente_normalizado = normalizar_id_cliente(id_cliente)
        fila_cliente = None
        for row in range(2, ws.max_row + 1):
            val_celda = normalizar_id_cliente(ws.cell(row=row, column=1).value)
            if val_celda == id_cliente_normalizado:
                fila_cliente = row
                break

        if not fila_cliente:
            print(f"❌ El ID de cliente '{id_cliente_normalizado}' no existe en reg_anon.xlsx.")
            wb.close()
            return False

        mes_pago_upper = str(mes_pago).strip().upper()
        if mes_pago_upper not in MESES:
            print(f"❌ El mes '{mes_pago_upper}' no es un mes cronológico válido.")
            wb.close()
            return False

        # Cascada de asignación: buscar siguiente mes vacío o con 0.0
        idx_mes_inicial = MESES.index(mes_pago_upper)
        mes_destino_final = None
        columna_destino_idx = None
        es_pago_adelantado = False

        for i in range(idx_mes_inicial, len(MESES)):
            mes_candidato = MESES[i]
            if mes_candidato in encabezados:
                col_idx = encabezados.index(mes_candidato) + 1
                valor_celda = ws.cell(row=fila_cliente, column=col_idx).value
                
                if valor_celda is None or str(valor_celda).strip() == "" or float(valor_celda or 0) == 0.0:
                    mes_destino_final = mes_candidato
                    columna_destino_idx = col_idx
                    if i > idx_mes_inicial:
                        es_pago_adelantado = True
                    break

        # Si todos los meses están llenos, acumular en el último
        if not mes_destino_final:
            print(f"⚠️ ¡Atención! Todos los meses posteriores están llenos. Se acumulará en el último mes registrado.")
            for i in reversed(range(len(MESES))):
                mes_candidato = MESES[i]
                if mes_candidato in encabezados:
                    mes_destino_final = mes_candidato
                    columna_destino_idx = encabezados.index(mes_candidato) + 1
                    es_pago_adelantado = True
                    break

        celda_objetivo = ws.cell(row=fila_cliente, column=columna_destino_idx)
        valor_actual = celda_objetivo.value
        
        try:
            monto_previo = float(valor_actual) if valor_actual is not None else 0.0
        except ValueError:
            monto_previo = 0.0

        nuevo_monto_total = monto_previo + monto_pago
        celda_objetivo.value = nuevo_monto_total

        # Aplicar estilos con openpyxl
        if es_pago_adelantado:
            print(f"🔄 El mes {mes_pago_upper} ya estaba liquidado. Redireccionando pago a {mes_destino_final}...")
            fuente_especial = Font(name='Arial', size=11, bold=True, italic=True, color='008000')
            celda_objetivo.font = fuente_especial
            print(f"   🎨 Aplicado formato visual especial (Negrita, Cursiva, Color Verde) en la celda.")
        else:
            celda_objetivo.font = Font(name='Arial', size=11, bold=False, italic=False, color='000000')

        wb.save(ARCHIVO_EXCEL_LOCAL)
        wb.close()
        print(f"📝 Registrado localmente en reg_anon.xlsx: {id_cliente_normalizado} -> ${nuevo_monto_total:.2f} en {mes_destino_final}")

        # 2. SINCRONIZACIÓN EN registros.xlsx
        id_real = None
        try:
            df_mapeo = pd.read_excel(ARCHIVO_MAPEO_LOCAL)
            df_mapeo.iloc[:, 0] = df_mapeo.iloc[:, 0].apply(normalizar_id_cliente)
            match_map = df_mapeo[df_mapeo.iloc[:, 0] == id_cliente_normalizado]
            if not match_map.empty:
                id_real = str(match_map.iloc[0, 1]).strip().upper()
                if id_real.endswith('.0'):
                    id_real = id_real[:-2]
        except Exception as e:
            print(f"⚠️ No se pudo consultar mapeo_identidad.xlsx para obtener el ID real: {e}")

        if id_real and os.path.exists(ARCHIVO_REGISTROS_LOCAL):
            try:
                wb_reg = openpyxl.load_workbook(ARCHIVO_REGISTROS_LOCAL)
                ws_reg = wb_reg.active
                
                encabezados_reg = [str(ws_reg.cell(row=1, column=col).value).strip().upper() for col in range(1, ws_reg.max_column + 1)]
                
                fila_real = None
                for row in range(2, ws_reg.max_row + 1):
                    val_celda = str(ws_reg.cell(row=row, column=1).value).strip().upper()
                    if val_celda.endswith('.0'):
                        val_celda = val_celda[:-2]
                    if val_celda == id_real:
                        fila_real = row
                        break

                if fila_real and mes_destino_final in encabezados_reg:
                    col_reg_idx = encabezados_reg.index(mes_destino_final) + 1
                    celda_reg = ws_reg.cell(row=fila_real, column=col_reg_idx)
                    
                    valor_actual_reg = celda_reg.value
                    try:
                        monto_previo_reg = float(valor_actual_reg) if valor_actual_reg is not None else 0.0
                    except ValueError:
                        monto_previo_reg = 0.0

                    nuevo_monto_reg = monto_previo_reg + monto_pago
                    celda_reg.value = nuevo_monto_reg
                    
                    if es_pago_adelantado:
                        celda_reg.font = Font(name='Arial', size=11, bold=True, italic=True, color='008000')
                    else:
                        celda_reg.font = Font(name='Arial', size=11, bold=False, italic=False, color='000000')

                    wb_reg.save(ARCHIVO_REGISTROS_LOCAL)
                    print(f"✅ Sincronizado localmente en registros.xlsx para el ID Real '{id_real}' en {mes_destino_final}: ${nuevo_monto_reg:.2f}")
                else:
                    if not fila_real:
                        print(f"⚠️ Advertencia: No se encontró la fila para el ID real '{id_real}' en registros.xlsx.")
                    else:
                        print(f"⚠️ Advertencia: El mes '{mes_destino_final}' no coincide con los encabezados de registros.xlsx.")
                
                wb_reg.close()
            except Exception as e:
                print(f"❌ Error al intentar escribir en registros.xlsx: {e}")
        else:
            if not id_real:
                print(f"⚠️ No se encontró ID real en el mapa de identidad para el cliente {id_cliente_normalizado}.")
            if not os.path.exists(ARCHIVO_REGISTROS_LOCAL):
                print(f"⚠️ No se encontró el archivo de registros en la ruta local: {ARCHIVO_REGISTROS_LOCAL}")

        return mes_destino_final
    except Exception as e:
        print(f"❌ Error al escribir estilos en los archivos Excel: {e}")
        return None

# ==========================================
# LÓGICA DE EJECUCIÓN PRINCIPAL
# ==========================================
def main():
    print("=== SISTEMA INTELIGENTE DE CONCILIACIÓN INSTITUCIONAL ===")
    
    if not API_KEY:
        print("⚠️ Advertencia: API_KEY no configurada. Asegúrate de declarar GEMINI_API_KEY en tu entorno.")

    # 1. Autenticar
    try:
        creds = obtener_credenciales()
        servicio_gmail = build('gmail', 'v1', credentials=creds)
        servicio_drive = build('drive', 'v3', credentials=creds)
    except Exception as e:
        print(f"❌ Error de autenticación de servicios de Google: {e}")
        return

    # 2. Descargar archivos necesarios de Google Drive de forma remota
    print("📂 Descargando archivos de Excel desde tu Google Drive...")
    id_mapeo = descargar_archivo_de_drive(servicio_drive, NOMBRE_MAPEO_DRIVE, ARCHIVO_MAPEO_LOCAL)
    id_excel = descargar_archivo_de_drive(servicio_drive, NOMBRE_EXCEL_DRIVE, ARCHIVO_EXCEL_LOCAL)
    id_registros = descargar_archivo_de_drive(servicio_drive, NOMBRE_REGISTROS_DRIVE, ARCHIVO_REGISTROS_LOCAL)

    if not id_mapeo or not id_excel or not id_registros:
        print("❌ Deteniendo el programa debido a la falta de archivos esenciales en Google Drive.")
        return

    # 3. Cargar mapeo de identidad
    try:
        df_mapeo = pd.read_excel(ARCHIVO_MAPEO_LOCAL)
        df_mapeo.iloc[:, 0] = df_mapeo.iloc[:, 0].apply(normalizar_id_cliente)
        df_mapeo.iloc[:, 3] = df_mapeo.iloc[:, 3].astype(str).str.strip().str.lower()
    except Exception as e:
        print(f"❌ Error al procesar mapeo_identidad.xlsx: {e}")
        return

    correos_clientes = df_mapeo.iloc[:, 3].dropna().unique().tolist()
    if not correos_clientes:
        print("⚠️ No hay cuentas de correo registradas en el mapa de identidad.")
        return

    # 4. Escanear Gmail
    transacciones_a_procesar = []
    
    label_id = obtener_o_crear_etiqueta(servicio_gmail, ETIQUETA_CONTROL)
    
    or_from_query = " OR ".join([f"from:{email}" for email in correos_clientes])
    query = f'has:attachment -label:{ETIQUETA_CONTROL} ({or_from_query})'
    print(f"🔍 Escaneando Gmail de forma inteligente con la consulta: {query}")

    try:
        resultado_busqueda = servicio_gmail.users().messages().list(userId='me', q=query).execute()
        mensajes = resultado_busqueda.get('messages', [])

        if not mensajes:
            print("📩 No se detectaron nuevos correos con archivos adjuntos de tus clientes.")
            return

        print(f"📩 Se detectaron {len(mensajes)} nuevos correos electrónicos para analizar.")

        for msg_info in mensajes:
            msg_id = msg_info['id']
            msg = servicio_gmail.users().messages().get(userId='me', id=msg_id).execute()
            thread_id = msg.get('threadId')
            payload = msg.get('payload', {})
            headers = payload.get('headers', [])

            from_header = next((h.get('value', '') for h in headers if h.get('name', '').lower() == 'from'), '')
            _, email_remitente = email.utils.parseaddr(from_header)
            email_remitente = email_remitente.strip().lower()

            match_cliente = df_mapeo[df_mapeo.iloc[:, 3] == email_remitente]
            if match_cliente.empty:
                continue

            client_id = match_cliente.iloc[0, 0]
            asunto_correo = next((h.get('value', '') for h in headers if h.get('name', '').lower() == 'subject'), "Sin Asunto")
            cuerpo_correo = extraer_cuerpo_texto(payload)

            if not buscar_palabras_clave(asunto_correo, cuerpo_correo):
                continue

            print(f"\n────────────────────────────────────────────────────────")
            print(f"👤 Remitente: {email_remitente} (Cliente: {client_id})")
            print(f"📧 Asunto: {asunto_correo}")

            # Detección del mes del pago
            mes_pago = detectar_mes_en_texto(asunto_correo, cuerpo_correo)
            if not mes_pago and thread_id:
                print(f"🔍 Mes no detectado en correo actual. Escaneando hilo de conversación...")
                mes_pago = detectar_mes_en_hilo(servicio_gmail, thread_id)

            adjuntos = obtener_adjuntos_recursivo(payload)
            for part in adjuntos:
                nombre_archivo = part.get('filename')
                if nombre_archivo:
                    ext = os.path.splitext(nombre_archivo)[1].lower()
                    mime_type = None
                    if ext in ['.jpg', '.jpeg']: mime_type = "image/jpeg"
                    elif ext == '.png': mime_type = "image/png"
                    elif ext == '.pdf': mime_type = "application/pdf"

                    if mime_type:
                        print(f"   📎 Procesando adjunto: {nombre_archivo}")
                        body = part.get('body', {})
                        attachment_id = body.get('attachmentId')

                        if attachment_id:
                            adjunto_raw = servicio_gmail.users().messages().attachments().get(
                                userId='me', messageId=msg_id, id=attachment_id).execute()
                            datos_bytes = base64.urlsafe_b64decode(adjunto_raw['data'].encode('UTF-8'))

                            resultado = extraer_metadatos_con_gemini(datos_bytes, mime_type)
                            if resultado and resultado.get('monto') is not None:
                                monto = float(resultado['monto'])
                                fecha_recibo = resultado.get('fecha', '')
                                print(f"      ✨ Extraído -> Fecha Recibo: {fecha_recibo} | Monto: ${monto:.2f}")

                                mes_final = mes_pago
                                if not mes_final:
                                    mes_final = obtener_mes_desde_fecha(fecha_recibo)
                                    if mes_final:
                                        print(f"      📄 Mes determinado por fecha del recibo: {mes_final}")

                                # Si sigue sin encontrarse el mes, usar fallback interactivo o automático de Cron
                                if not mes_final:
                                    if sys.stdin.isatty():
                                        while True:
                                            entrada = input(f"      👉 Introduce el mes para este pago (ej. MARZO) o ENTER para omitir: ").strip().upper()
                                            if not entrada:
                                                print("      ⏭️ Correo omitido por el usuario.")
                                                break
                                            if entrada in MESES:
                                                mes_final = entrada
                                                break
                                            else:
                                                print(f"      ⚠️ '{entrada}' no es un mes válido. Inténtalo de nuevo.")
                                    else:
                                        # Cron / Ejecución desatendida: Asigna el mes actual del sistema
                                        mes_final = MESES[datetime.datetime.now().month - 1]
                                        print(f"      🤖 Ejecución no interactiva (Cron/VPS). Asignando mes en curso: {mes_final}")

                                if mes_final:
                                    transacciones_a_procesar.append({
                                        "client_id": client_id,
                                        "mes": mes_final,
                                        "monto": monto,
                                        "msg_id": msg_id,
                                        "label_id": label_id,
                                        "remitente": email_remitente
                                    })
                            else:
                                print(f"      ⚠️ No se pudo extraer información estructurada del adjunto.")

    except Exception as e:
        print(f"❌ Error durante el análisis de Gmail: {e}")
        return

    # 5. Escribir registros en Excel y sincronizar en Google Drive
    if transacciones_a_procesar:
        print("\n=== INICIANDO ACTUALIZACIÓN DE REGISTROS ===")
        cambios_realizados = False
        
        for tx in transacciones_a_procesar:
            # Escribir en los excel locales, aplicar formato y cruzar identidades
            mes_registrado = actualizar_excel(tx["client_id"], tx["mes"], tx["monto"])
            
            if mes_registrado:
                cambios_realizados = True
                # Marcar correo analizado como procesado
                aplicar_etiqueta_procesado(servicio_gmail, tx["msg_id"], tx["label_id"])
            else:
                print(f"⚠️ Se omitió el etiquetado del correo {tx['msg_id']} debido a un error al guardar en Excel.")

        # 6. Sincronizar de regreso a Drive únicamente si hubo cambios reales
        if cambios_realizados:
            actualizar_archivo_en_drive(servicio_drive, id_excel, ARCHIVO_EXCEL_LOCAL)
            actualizar_archivo_en_drive(servicio_drive, id_registros, ARCHIVO_REGISTROS_LOCAL)
    else:
        print("\n❌ No se detectaron nuevos pagos válidos de clientes en esta ejecución.")

if __name__ == "__main__":
    main()
